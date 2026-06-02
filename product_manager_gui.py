import sys
import sqlite3
from typing import List, Tuple, Optional
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QTableWidget, QTableWidgetItem,
    QMessageBox, QSpinBox, QHeaderView
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class ProductDatabase:
    """SQLite를 사용한 제품 데이터베이스 관리 클래스"""
    
    def __init__(self, db_name: str = "products.db"):
        """데이터베이스 연결 초기화"""
        self.conn = sqlite3.connect(db_name)
        self.cursor = self.conn.cursor()
        self.create_table()
    
    def create_table(self):
        """Products 테이블 생성"""
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS Products (
                productID INTEGER PRIMARY KEY AUTOINCREMENT,
                productName TEXT NOT NULL,
                productPrice INTEGER NOT NULL
            )
        ''')
        self.conn.commit()
    
    def insert_product(self, product_name: str, product_price: int) -> bool:
        """제품 입력"""
        try:
            self.cursor.execute('''
                INSERT INTO Products (productName, productPrice)
                VALUES (?, ?)
            ''', (product_name, product_price))
            self.conn.commit()
            return True
        except sqlite3.Error as e:
            print(f"✗ 입력 오류: {e}")
            return False
    
    def get_all_products(self) -> List[Tuple]:
        """모든 제품 조회"""
        try:
            self.cursor.execute('SELECT * FROM Products')
            products = self.cursor.fetchall()
            return products
        except sqlite3.Error as e:
            print(f"✗ 검색 오류: {e}")
            return []
    
    def update_product(self, product_id: int, product_name: str = None, 
                      product_price: int = None) -> bool:
        """제품 수정"""
        try:
            if product_name and product_price:
                self.cursor.execute('''
                    UPDATE Products 
                    SET productName = ?, productPrice = ?
                    WHERE productID = ?
                ''', (product_name, product_price, product_id))
            elif product_name:
                self.cursor.execute('''
                    UPDATE Products 
                    SET productName = ?
                    WHERE productID = ?
                ''', (product_name, product_id))
            elif product_price:
                self.cursor.execute('''
                    UPDATE Products 
                    SET productPrice = ?
                    WHERE productID = ?
                ''', (product_price, product_id))
            self.conn.commit()
            return True
        except sqlite3.Error as e:
            print(f"✗ 수정 오류: {e}")
            return False
    
    def delete_product(self, product_id: int) -> bool:
        """제품 삭제"""
        try:
            self.cursor.execute('DELETE FROM Products WHERE productID = ?', 
                              (product_id,))
            self.conn.commit()
            return True
        except sqlite3.Error as e:
            print(f"✗ 삭제 오류: {e}")
            return False
    
    def close(self):
        """데이터베이스 연결 종료"""
        self.conn.close()


class ProductManagerGUI(QMainWindow):
    """제품 관리 GUI 애플리케이션"""
    
    def __init__(self):
        super().__init__()
        self.db = ProductDatabase("products.db")
        self.current_edit_id = None
        self.init_ui()
        self.apply_styles()
    
    def init_ui(self):
        """UI 초기화"""
        self.setWindowTitle("제품 관리 시스템")
        self.setGeometry(100, 100, 800, 600)
        
        # 중앙 위젯
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout()
        
        # ===== 상단: 입력 영역 =====
        input_layout = QHBoxLayout()
        
        # 제품명 입력
        input_layout.addWidget(QLabel("상품명:"))
        self.product_name_input = QLineEdit()
        self.product_name_input.setPlaceholderText("제품명을 입력하세요")
        input_layout.addWidget(self.product_name_input)
        
        # 가격 입력
        input_layout.addWidget(QLabel("가격:"))
        self.product_price_input = QSpinBox()
        self.product_price_input.setMaximum(9999999)
        self.product_price_input.setSuffix(" 원")
        input_layout.addWidget(self.product_price_input)
        
        # 추가 버튼
        self.add_button = QPushButton("추가/수정")
        self.add_button.clicked.connect(self.add_product)
        input_layout.addWidget(self.add_button)
        
        main_layout.addLayout(input_layout)
        
        # ===== 중간: 데이터 리스트 (테이블) =====
        list_label = QLabel("제품 목록:")
        main_layout.addWidget(list_label)
        
        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["ID", "상품명", "가격(원)"])
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.setSelectionBehavior(self.table.SelectRows)
        self.table.setSelectionMode(self.table.SingleSelection)
        self.table.doubleClicked.connect(self.on_table_double_click)
        main_layout.addWidget(self.table)
        
        # ===== 하단: 버튼 영역 =====
        button_layout = QHBoxLayout()
        
        # 초기화 버튼
        self.clear_button = QPushButton("초기화")
        self.clear_button.clicked.connect(self.clear_input)
        button_layout.addWidget(self.clear_button)
        
        # 삭제 버튼
        self.delete_button = QPushButton("삭제")
        self.delete_button.clicked.connect(self.delete_product)
        button_layout.addWidget(self.delete_button)
        
        # 새로고침 버튼
        self.refresh_button = QPushButton("새로고침")
        self.refresh_button.clicked.connect(self.load_products)
        button_layout.addWidget(self.refresh_button)
        
        button_layout.addStretch()
        
        # 엑셀 저장 버튼
        self.export_button = QPushButton("엑셀로 저장")
        self.export_button.clicked.connect(self.export_to_excel)
        button_layout.addWidget(self.export_button)
        
        main_layout.addLayout(button_layout)
        
        central_widget.setLayout(main_layout)
        
        # 초기 데이터 로드
        self.load_products()
    
    def apply_styles(self):
        """스타일시트 적용"""
        style = """
        QMainWindow {
            background-color: #f0f2f5;
        }
        QLabel {
            font-size: 11pt;
            font-weight: bold;
            color: #333333;
        }
        QLineEdit {
            border: 2px solid #ddd;
            border-radius: 5px;
            padding: 8px;
            font-size: 11pt;
            background-color: white;
            selection-background-color: #4CAF50;
        }
        QLineEdit:focus {
            border: 2px solid #4CAF50;
            background-color: #f9fff9;
        }
        QSpinBox {
            border: 2px solid #ddd;
            border-radius: 5px;
            padding: 5px;
            font-size: 11pt;
            background-color: white;
        }
        QSpinBox:focus {
            border: 2px solid #4CAF50;
        }
        QPushButton {
            border: none;
            border-radius: 5px;
            padding: 10px 20px;
            font-size: 10pt;
            font-weight: bold;
            color: white;
            transition: all 0.3s;
        }
        QPushButton#addEditBtn {
            background-color: #4CAF50;
        }
        QPushButton#addEditBtn:hover {
            background-color: #45a049;
        }
        QPushButton#addEditBtn:pressed {
            background-color: #3d8b40;
        }
        QPushButton#clearBtn {
            background-color: #ff9800;
        }
        QPushButton#clearBtn:hover {
            background-color: #e68900;
        }
        QPushButton#deleteBtn {
            background-color: #f44336;
        }
        QPushButton#deleteBtn:hover {
            background-color: #da190b;
        }
        QPushButton#refreshBtn {
            background-color: #2196F3;
        }
        QPushButton#refreshBtn:hover {
            background-color: #0b7dda;
        }
        QPushButton#exportBtn {
            background-color: #FF6B6B;
        }
        QPushButton#exportBtn:hover {
            background-color: #ff5252;
        }
        QTableWidget {
            border: 1px solid #ddd;
            border-radius: 5px;
            gridline-color: #e0e0e0;
            background-color: white;
        }
        QTableWidget::item {
            padding: 5px;
            border-bottom: 1px solid #f0f0f0;
        }
        QTableWidget::item:selected {
            background-color: #4CAF50;
            color: white;
        }
        QHeaderView::section {
            background-color: #1976D2;
            color: white;
            padding: 5px;
            border: none;
            font-weight: bold;
            font-size: 11pt;
        }
        """
        self.setStyleSheet(style)
        
        # 각 버튼에 ObjectName 설정하여 스타일 적용
        self.add_button.setObjectName("addEditBtn")
        self.clear_button.setObjectName("clearBtn")
        self.delete_button.setObjectName("deleteBtn")
        self.refresh_button.setObjectName("refreshBtn")
        self.export_button.setObjectName("exportBtn")
    
    def add_product(self):
        """제품 추가 또는 수정"""
        product_name = self.product_name_input.text().strip()
        product_price = self.product_price_input.value()
        
        if not product_name:
            QMessageBox.warning(self, "경고", "상품명을 입력하세요.")
            return
        
        if product_price <= 0:
            QMessageBox.warning(self, "경고", "가격을 0원 이상으로 입력하세요.")
            return
        
        # 현재 편집 중인 상품이 있으면 수정, 없으면 추가
        if self.current_edit_id is not None:
            if self.db.update_product(self.current_edit_id, product_name, product_price):
                QMessageBox.information(self, "성공", f"'{product_name}'이(가) 수정되었습니다.")
                self.clear_input()
                self.load_products()
            else:
                QMessageBox.critical(self, "오류", "제품 수정에 실패했습니다.")
        else:
            if self.db.insert_product(product_name, product_price):
                QMessageBox.information(self, "성공", f"'{product_name}'이(가) 추가되었습니다.")
                self.clear_input()
                self.load_products()
            else:
                QMessageBox.critical(self, "오류", "제품 추가에 실패했습니다.")
    
    def load_products(self):
        """제품 목록 로드"""
        self.table.setRowCount(0)
        products = self.db.get_all_products()
        
        for row, product in enumerate(products):
            self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(str(product[0])))
            self.table.setItem(row, 1, QTableWidgetItem(product[1]))
            self.table.setItem(row, 2, QTableWidgetItem(f"{product[2]:,}"))
    
    def on_table_double_click(self, index):
        """테이블 더블클릭 이벤트 - 선택된 행의 데이터를 입력 필드에 복사"""
        current_row = self.table.currentRow()
        if current_row < 0:
            return
        
        # 테이블에서 데이터 추출
        product_id = int(self.table.item(current_row, 0).text())
        product_name = self.table.item(current_row, 1).text()
        product_price = int(self.table.item(current_row, 2).text().replace(",", ""))
        
        # 입력 필드에 데이터 채우기
        self.current_edit_id = product_id
        self.product_name_input.setText(product_name)
        self.product_price_input.setValue(product_price)
        
        # 상단으로 스크롤 (입력 필드가 보이도록)
        self.product_name_input.setFocus()
        QMessageBox.information(
            self, "수정 모드",
            f"ID {product_id}의 '{product_name}' 상품을 수정하고 있습니다.\n수정 후 '추가/수정' 버튼을 클릭하세요.\n초기화 버튼으로 취소할 수 있습니다."
        )
    
    def clear_input(self):
        """입력 필드 초기화"""
        self.current_edit_id = None
        self.product_name_input.clear()
        self.product_price_input.setValue(0)
        self.product_name_input.setFocus()
    
    def delete_product(self):
        """제품 삭제"""
        # 편집 중인 상품이 있으면 그것을 삭제, 없으면 테이블 선택 사용
        if self.current_edit_id is not None:
            product_id = self.current_edit_id
            product_name = self.product_name_input.text().strip()
        else:
            current_row = self.table.currentRow()
            if current_row < 0:
                QMessageBox.warning(self, "경고", "삭제할 제품을 선택하세요.")
                return
            
            product_id = int(self.table.item(current_row, 0).text())
            product_name = self.table.item(current_row, 1).text()
        
        reply = QMessageBox.question(
            self, "삭제 확인",
            f"'{product_name}'(ID: {product_id})을(를) 삭제하시겠습니까?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            if self.db.delete_product(product_id):
                QMessageBox.information(self, "성공", f"'{product_name}'이(가) 삭제되었습니다.")
                self.clear_input()
                self.load_products()
            else:
                QMessageBox.critical(self, "오류", "제품 삭제에 실패했습니다.")
    
    def export_to_excel(self):
        """데이터를 엑셀 파일로 저장"""
        products = self.db.get_all_products()
        
        if not products:
            QMessageBox.warning(self, "경고", "저장할 데이터가 없습니다.")
            return
        
        try:
            # 워크북 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "제품"
            
            # 헤더 스타일 설정
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 헤더 작성
            headers = ["ID", "상품명", "가격(원)"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 데이터 작성
            for row, product in enumerate(products, start=2):
                ws.cell(row=row, column=1).value = product[0]
                ws.cell(row=row, column=2).value = product[1]
                ws.cell(row=row, column=3).value = product[2]
                
                # 가격 열 오른쪽 정렬
                ws.cell(row=row, column=3).alignment = Alignment(horizontal="right")
            
            # 열 너비 조정
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 15
            
            # 파일 저장
            filename = "products.xlsx"
            wb.save(filename)
            QMessageBox.information(
                self, "성공",
                f"데이터가 '{filename}'으로 저장되었습니다.\n총 {len(products)}개의 제품이 저장되었습니다."
            )
        except Exception as e:
            QMessageBox.critical(self, "오류", f"엑셀 저장 중 오류가 발생했습니다:\n{str(e)}")
    
    def closeEvent(self, event):
        """애플리케이션 종료 시"""
        self.db.close()
        event.accept()


def main():
    app = QApplication(sys.argv)
    window = ProductManagerGUI()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
